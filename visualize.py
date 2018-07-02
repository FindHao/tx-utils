# 读写2003 excel
import argparse
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.layout import Layout, ManualLayout

import re

# 自从jetpack3.2以后，tegrastats功能更强大了。
tegrastats_version = 3.2


class Status:
    def __init__(self):
        # CPU部分
        # 每个cpu核心的频率变化
        self.cpu_freq = [[], [], [], [], [], []]
        # 占用率变化
        self.cpu_utilization = [[], [], [], [], [], []]
        # GPU部分
        self.ram = []
        self.emc = []
        self.gpu = []


def filter_content_from_raw_log(start_time, end_time, file_path):
    """将频率数据从原始log中提取出来"""
    with open(file_path, 'r') as fin:
        content = fin.read()
    print("start time:\t%s" % start_time)
    print("end time:\t%s" % end_time)
    if not start_time:
        start_time = ''
    if not end_time:
        end_time = ''
    reg = re.compile(r'%s[\s\S]+%s' % (start_time, end_time))
    if tegrastats_version >= 3.2:
        # fix bug that when some cores are closed the reg doesn't work
        reg2 = re.compile(
            r'RAM (\d+).+?CPU \[(.+?)\] EMC_FREQ (\d+)%@.+?GR3D_FREQ (\d+)%')
    else:
        reg2 = re.compile(
            r'RAM (\d+).+?cpu \[(.+?)\] EMC (\d+)%@.+?GR3D (\d+)%')
    reg_cpu = re.compile(r'(\d+)\%@(\d+)')
    result = reg.findall(content)
    status = Status()
    if result:
        result2 = reg2.findall(result[0])
        if result2:
            for line in result2:
                status.ram.append(int(line[0]))
                # cpu status
                temp_cpus = line[1].split(',')
                for i, a_cpu in enumerate(temp_cpus):
                    cpu_result = reg_cpu.findall(a_cpu)
                    if cpu_result:
                        cpu_result = cpu_result[0]
                        status.cpu_utilization[i].append(int(cpu_result[0]))
                        status.cpu_freq[i].append(int(cpu_result[1]))
                    else:
                        # this core is closed.
                        status.cpu_utilization[i].append(0)
                        status.cpu_freq[i].append(0)
                status.emc.append(int(line[2]))
                status.gpu.append(int(line[3]))
        else:
            print("没有符合条件的输出-2")
    else:
        print("没有符合条件的输出-1")
    return status


def write_status_to_xls(status, xls_file='./log.xls'):
    """将获取的内容写到excel文件里"""
    wb = Workbook()
    sheet_name = 'cpu'
    sheet = wb.active
    sheet.title = sheet_name
    cpu_status = ['cpu0', 'cpu1', 'cpu2', 'cpu3', 'cpu4', 'cpu5']
    # 真正的数据从第三行开始
    base_offset_value = 3
    # cpu频率
    sheet.cell(row=1, column=1, value='cpu frequency')
    for i in range(6):
        sheet.cell(row=2, column=i + 1, value=cpu_status[i])
    for i in range(0, len(status.cpu_freq[0])):
        for j in range(0, 6):
            sheet.cell(row=i + base_offset_value, column=j + 1, value=status.cpu_freq[j][i])
    # cpu占用率
    base_offset_cpu_utilization = 8
    sheet.cell(row=1, column=base_offset_cpu_utilization, value='cpu utilization')
    for i in range(6):
        sheet.cell(row=2, column=i + base_offset_cpu_utilization, value=cpu_status[i])
    for i in range(0, len(status.cpu_utilization[0])):
        for j in range(0, 6):
            sheet.cell(row=i + base_offset_value, column=j + base_offset_cpu_utilization,
                       value=status.cpu_utilization[j][i])
    # ram
    base_offset_ram = 15
    sheet.cell(row=2, column=base_offset_ram, value='ram')
    for i in range(0, len(status.ram)):
        sheet.cell(row=i + base_offset_value, column=base_offset_ram, value=status.ram[i])
    # emc
    base_offset_emc = 16
    sheet.cell(row=2, column=base_offset_emc, value='emc')
    for i in range(0, len(status.emc)):
        sheet.cell(row=i + base_offset_value, column=base_offset_emc, value=status.emc[i])
    # gpu
    base_offset_gpu = 17
    sheet.cell(row=2, column=base_offset_gpu, value='gpu')
    for i in range(0, len(status.gpu)):
        sheet.cell(row=i + base_offset_value, column=base_offset_gpu, value=status.gpu[i])

    # 制作line chart
    chart = LineChart()
    chart.y_axis.title = 'Rate'
    # where is the data
    data = Reference(sheet, min_col=base_offset_cpu_utilization, min_row=2,
                     max_row=len(status.gpu) + base_offset_value - 1, max_col=base_offset_cpu_utilization + 5)
    data_gpu = Reference(sheet, min_col=base_offset_gpu, min_row=2, max_row=len(status.gpu) + base_offset_value - 1,
                         max_col=base_offset_gpu)
    chart.add_data(data, titles_from_data=True)
    chart.add_data(data_gpu, titles_from_data=True)
    # chart.layout = Layout(
    #     manualLayout=ManualLayout(
    #         x=2, y=2,
    #         h=2, w=2,
    #     )
    # )
    # chart.add_series({
    #     # 'categories': '=%s!$H$3:$H$%d' % (sheet_name, 2+len(status.gpu)),
    #     'values': '=%s!$H$3:$H$%d' % (sheet_name, 2+len(status.gpu)),
    #     'line': {'color': 'red'},
    # }
    # )
    sheet.add_chart(chart, 'A%d' % (len(status.gpu) + base_offset_value + 1))

    wb.save(xls_file)
    print("Done!")


if __name__ == '__main__':
    # 输入参数
    parser = argparse.ArgumentParser(description='用来提取一段时间内tegra cpu gpu利用率等信息的工具')
    parser.add_argument('-s', '--start', metavar='开始的时间戳', required=False, dest='start_time', action='store')
    parser.add_argument('-e', '--end', metavar='结束的时间', required=False, dest='end_time', action='store')
    parser.add_argument('-i', '--input', metavar='输入的log文件', required=True, dest='log_file_path', action='store')
    parser.add_argument('-o', '--output', metavar='输出的excel文件路径', required=False, dest='xls_file_path', action='store')
    args = parser.parse_args()
    log_file = args.log_file_path
    start_time = args.start_time
    end_time = args.end_time
    # 默认生成的xls表格是当前目录下的log.xls
    xls_file = args.xls_file_path
    status = filter_content_from_raw_log(start_time, end_time, log_file)
    write_status_to_xls(status, xls_file)
